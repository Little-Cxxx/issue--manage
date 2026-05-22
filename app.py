# -*- coding: utf-8 -*-
import os
import sqlite3
import io
import zipfile
import tempfile
import shutil
from datetime import datetime, date
from flask import Flask, request, jsonify, send_file, render_template
from werkzeug.utils import secure_filename
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB
app.config['ADMIN_PASSWORD'] = 'admin123'

DB_PATH = 'issues.db'

# ============ Excel 列名映射 ============
COLUMNS = [
    '测试事项', '阶段', '分类', '测试工程师', '研发工程师', '部门',
    '发生日期', '要求结案日期', '统计日期', '是否Delay', 'Delay天数',
    'Issue状态', '严重度', '问题点详细描述', 'RD是否提供改善措施',
    '要求回复日期', '回复是否Delay', '改善措施', 'DQA是否确认', 'DQA确认',
    '问题点当前处理进度', '更新时间'
]

STATUS_OPTIONS = ['待确认', '已确认', '处理中', '待回归', '已回归', '待关闭']

# Excel 导出时状态列颜色映射（ARGB格式）
STATUS_COLORS = {
    '待确认': 'FFFF0000',
    '已确认': 'FFFFC000',
    '处理中': 'FFC6EFCE',
    '待回归': 'FF00B050',
    '已回归': 'FF4472C4',
    '待关闭': 'FF203864',
}


# ============ 数据库初始化 ============
def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    conn = get_db()
    cursor = conn.cursor()

    # 创建项目表
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            created_at TEXT NOT NULL
        )
    ''')

    # 检查 issues 表是否已有 project_id 列
    table_exists = cursor.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name='issues'"
    ).fetchone()
    existing_cols = []
    if table_exists:
        cursor.execute("PRAGMA table_info(issues)")
        existing_cols = [row[1] for row in cursor.fetchall()]

    if not table_exists:
        # 全新创建
        cols_def = ', '.join([f'"{col}" TEXT' for col in COLUMNS])
        cursor.execute(f'''
            CREATE TABLE IF NOT EXISTS issues (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL DEFAULT 1,
                {cols_def}
            )
        ''')
    elif 'project_id' not in existing_cols:
        # 旧表需要迁移：添加 project_id 列
        cols_def = ', '.join([f'"{col}" TEXT' for col in COLUMNS])
        cursor.execute(f'''
            CREATE TABLE issues_new (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                project_id INTEGER NOT NULL DEFAULT 1,
                {cols_def}
            )
        ''')
        col_names = ', '.join([f'"{c}"' for c in COLUMNS])
        cursor.execute(f'INSERT INTO issues_new (project_id, {col_names}) SELECT 1, {col_names} FROM issues')
        cursor.execute('DROP TABLE issues')
        cursor.execute('ALTER TABLE issues_new RENAME TO issues')

    # 确保有默认项目
    cursor.execute('SELECT COUNT(*) FROM projects')
    if cursor.fetchone()[0] == 0:
        cursor.execute("INSERT INTO projects (name, created_at) VALUES (?, ?)",
                       ('默认项目', date.today().strftime('%Y-%m-%d')))

    conn.commit()
    conn.close()


# ============ 辅助函数 ============
def parse_excel(filepath):
    """解析 Excel 文件，返回数据列表"""
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    rows = []
    headers = None

    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h).strip() if h else '' for h in row]
            continue
        row_data = {}
        for i, header in enumerate(headers):
            if header in COLUMNS:
                val = row[i] if i < len(row) else ''
                if isinstance(val, datetime):
                    val = val.strftime('%Y-%m-%d %H:%M:%S')
                else:
                    val = str(val).strip() if val is not None else ''
                row_data[header] = val
        if row_data:
            rows.append(row_data)

    wb.close()
    return rows


def export_excel(project_id):
    """导出指定项目数据为 Excel（含标题格式和状态颜色）"""
    conn = get_db()
    cursor = conn.cursor()
    cols_str = ', '.join([f'"{c}"' for c in COLUMNS])
    cursor.execute(f'SELECT {cols_str} FROM issues WHERE project_id = ?', (project_id,))
    rows = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Issue跟踪'

    # 标题行样式：蓝色背景 + 白色加粗字体
    header_font = Font(name='微软雅黑', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # 状态列索引（从1开始）
    status_col_idx = COLUMNS.index('问题点当前处理进度') + 1

    for row_idx, row in enumerate(rows, 2):
        for col_idx, col_name in enumerate(COLUMNS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row[col_name])
            # 状态列添加背景色
            if col_idx == status_col_idx:
                status_val = row[col_name] or ''
                color = STATUS_COLORS.get(status_val)
                if color:
                    cell.fill = PatternFill(start_color=color[2:], end_color=color[2:], fill_type='solid')
                    # 深色背景用白字，浅色背景用黑字
                    if status_val in ('待确认', '待回归', '已回归', '待关闭'):
                        cell.font = Font(color='FFFFFF')
                    else:
                        cell.font = Font(color='000000')

    for col_idx in range(1, len(COLUMNS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

    # 冻结首行
    ws.freeze_panes = 'A2'

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    wb.close()
    return output


def extract_zip(zip_path):
    """从 .zip 压缩包中提取 Excel 文件，返回 Excel 文件路径"""
    extract_dir = tempfile.mkdtemp()
    try:
        with zipfile.ZipFile(zip_path, 'r') as zf:
            excel_files = [f for f in zf.namelist()
                          if f.lower().endswith(('.xlsx', '.xls'))
                          and not f.startswith('__MACOSX')
                          and not os.path.basename(f).startswith('~$')]
            if not excel_files:
                raise ValueError('ZIP 压缩包中未找到 .xlsx 或 .xls 文件')

            zf.extractall(extract_dir)
            excel_path = os.path.join(extract_dir, excel_files[0])
            return excel_path, extract_dir
    except zipfile.BadZipFile:
        raise ValueError('无法解析 ZIP 文件，请确认文件格式正确')
    except Exception:
        shutil.rmtree(extract_dir, ignore_errors=True)
        raise


# ============ 路由 ============
@app.route('/')
def index():
    return render_template('index.html')


# ============ 项目管理 API ============
@app.route('/api/projects', methods=['GET'])
def get_projects():
    """获取所有项目列表"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT id, name, created_at FROM projects ORDER BY id')
    rows = [dict(r) for r in cursor.fetchall()]
    conn.close()
    return jsonify({'success': True, 'data': rows})


@app.route('/api/projects', methods=['POST'])
def create_project():
    """创建新项目"""
    data = request.get_json()
    name = data.get('name', '').strip()
    if not name:
        return jsonify({'success': False, 'message': '项目名称不能为空'})

    conn = get_db()
    cursor = conn.cursor()
    now = date.today().strftime('%Y-%m-%d')
    cursor.execute('INSERT INTO projects (name, created_at) VALUES (?, ?)', (name, now))
    conn.commit()
    pid = cursor.lastrowid
    conn.close()
    return jsonify({'success': True, 'message': f'项目「{name}」创建成功', 'id': pid})


@app.route('/api/projects/<int:project_id>', methods=['DELETE'])
def delete_project(project_id):
    """删除项目及其所有数据"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('SELECT COUNT(*) FROM projects')
    if cursor.fetchone()[0] <= 1:
        conn.close()
        return jsonify({'success': False, 'message': '至少需要保留一个项目'})

    cursor.execute('DELETE FROM issues WHERE project_id = ?', (project_id,))
    cursor.execute('DELETE FROM projects WHERE id = ?', (project_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': '项目已删除'})


@app.route('/api/projects/<int:project_id>/clear', methods=['DELETE'])
def clear_project_issues(project_id):
    """一键清空指定项目的所有 Issue 数据"""
    conn = get_db()
    cursor = conn.cursor()
    cursor.execute('DELETE FROM issues WHERE project_id = ?', (project_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'message': '当前表数据已清空'})


# ============ Issue API (均带 project_id) ============
@app.route('/api/upload', methods=['POST'])
def upload_excel():
    """上传 Excel 或 ZIP 并导入数据库"""
    if 'file' not in request.files:
        return jsonify({'success': False, 'message': '未选择文件'})

    file = request.files['file']
    if file.filename == '':
        return jsonify({'success': False, 'message': '未选择文件'})

    fname_lower = file.filename.lower()
    if not (fname_lower.endswith(('.xlsx', '.xls', '.zip'))):
        return jsonify({'success': False, 'message': '仅支持 .xlsx、.xls 或 .zip 格式'})

    project_id = request.form.get('project_id', '1')

    filename = secure_filename(file.filename)
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)

    extract_dir = None
    try:
        if fname_lower.endswith('.zip'):
            excel_path, extract_dir = extract_zip(filepath)
        else:
            excel_path = filepath

        data = parse_excel(excel_path)
        if not data:
            return jsonify({'success': False, 'message': '文件中没有有效数据'})

        conn = get_db()
        cursor = conn.cursor()
        # 覆盖模式：清空当前项目数据
        cursor.execute('DELETE FROM issues WHERE project_id = ?', (project_id,))

        for row in data:
            columns = list(row.keys())
            placeholders = ', '.join(['?' for _ in columns])
            values = [project_id] + [row[c] for c in columns]
            cols_str = ', '.join(['project_id'] + [f'"{c}"' for c in columns])
            cursor.execute(f'INSERT INTO issues ({cols_str}) VALUES ({placeholders})', values)

        conn.commit()
        conn.close()

        return jsonify({
            'success': True,
            'message': f'成功导入 {len(data)} 条记录',
            'count': len(data)
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'解析失败: {str(e)}'})
    finally:
        if os.path.exists(filepath):
            os.remove(filepath)
        if extract_dir and os.path.exists(extract_dir):
            shutil.rmtree(extract_dir, ignore_errors=True)


@app.route('/api/issues/bulk-save', methods=['POST'])
def bulk_save_issues():
    """批量保存 Issue 数据（在线粘贴录入）"""
    data = request.get_json()
    rows = data.get('rows', [])
    append = data.get('append', True)
    project_id = data.get('project_id', '1')

    if not rows and not append:
        # 允许空 rows + 非追加模式（即清空操作）
        conn = get_db()
        cursor = conn.cursor()
        cursor.execute('DELETE FROM issues WHERE project_id = ?', (project_id,))
        conn.commit()
        conn.close()
        return jsonify({'success': True, 'message': '数据已清空', 'count': 0})

    if not rows:
        return jsonify({'success': False, 'message': '没有要保存的数据'})

    conn = get_db()
    cursor = conn.cursor()

    try:
        if not append:
            cursor.execute('DELETE FROM issues WHERE project_id = ?', (project_id,))

        saved = 0
        for row in rows:
            valid_row = {k: v for k, v in row.items() if k in COLUMNS}
            if not valid_row:
                continue
            columns = ['project_id'] + list(valid_row.keys())
            placeholders = ', '.join(['?' for _ in columns])
            values = [project_id] + [valid_row[k] for k in valid_row.keys()]
            cols_str = ', '.join([f'"{c}"' for c in columns])
            cursor.execute(f'INSERT INTO issues ({cols_str}) VALUES ({placeholders})', values)
            saved += 1

        conn.commit()
        return jsonify({
            'success': True,
            'message': f'成功保存 {saved} 条记录',
            'count': saved
        })
    except Exception as e:
        conn.rollback()
        return jsonify({'success': False, 'message': f'保存失败: {str(e)}'})
    finally:
        conn.close()


@app.route('/api/issues', methods=['GET'])
def get_issues():
    """获取 Issue 列表（按项目+工程师筛选）"""
    engineer = request.args.get('engineer', '').strip()
    is_admin = request.args.get('admin', 'false').lower() == 'true'
    project_id = request.args.get('project_id', '').strip()

    conn = get_db()
    cursor = conn.cursor()

    if is_admin:
        # 管理员按指定项目查询
        if project_id:
            cursor.execute('SELECT * FROM issues WHERE project_id = ? ORDER BY id', (project_id,))
        else:
            cursor.execute('SELECT * FROM issues ORDER BY id')
    elif engineer:
        # 普通用户：查询该工程师在所有项目中的 Issue
        if project_id:
            cursor.execute(
                'SELECT * FROM issues WHERE project_id = ? AND "研发工程师" = ? ORDER BY id',
                (project_id, engineer))
        else:
            cursor.execute(
                'SELECT * FROM issues WHERE "研发工程师" = ? ORDER BY id',
                (engineer,))
    else:
        conn.close()
        return jsonify({'success': False, 'message': '请提供工程师姓名'})

    rows = cursor.fetchall()
    conn.close()

    issues = [{key: row[key] for key in row.keys()} for row in rows]
    return jsonify({'success': True, 'data': issues, 'count': len(issues)})


@app.route('/api/issues/<int:issue_id>', methods=['PUT'])
def update_issue(issue_id):
    """更新 Issue 的问题点当前处理进度和更新时间"""
    data = request.get_json()
    status = data.get('status', '').strip()
    update_time = data.get('update_time', '').strip()

    if status and status not in STATUS_OPTIONS:
        return jsonify({
            'success': False,
            'message': f'状态值无效，可选: {", ".join(STATUS_OPTIONS)}'
        })

    if not status and not update_time:
        return jsonify({'success': False, 'message': '未提供更新内容'})

    conn = get_db()
    cursor = conn.cursor()

    updates = []
    values = []

    if status:
        updates.append('"问题点当前处理进度" = ?')
        values.append(status)
    if update_time:
        updates.append('"更新时间" = ?')
        values.append(update_time)

    values.append(issue_id)
    sql = f'UPDATE issues SET {", ".join(updates)} WHERE id = ?'
    cursor.execute(sql, values)
    conn.commit()
    conn.close()

    return jsonify({'success': True, 'message': '更新成功'})


@app.route('/api/export', methods=['GET'])
def export_issues():
    """管理员导出 Excel"""
    password = request.args.get('password', '')
    project_id = request.args.get('project_id', '1')

    if password != app.config['ADMIN_PASSWORD']:
        return jsonify({'success': False, 'message': '管理员密码错误'})

    try:
        output = export_excel(project_id)
        timestamp = date.today().strftime('%Y%m%d')
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Issue跟踪_{timestamp}.xlsx'
        )
    except Exception as e:
        return jsonify({'success': False, 'message': f'导出失败: {str(e)}'})


@app.route('/api/verify-admin', methods=['POST'])
def verify_admin():
    """验证管理员密码"""
    data = request.get_json()
    password = data.get('password', '')
    if password == app.config['ADMIN_PASSWORD']:
        return jsonify({'success': True})
    return jsonify({'success': False, 'message': '管理员密码错误'})


@app.route('/api/stats', methods=['GET'])
def get_stats():
    """获取统计信息（按项目）"""
    project_id = request.args.get('project_id', '1')

    conn = get_db()
    cursor = conn.cursor()

    cursor.execute('SELECT COUNT(*) FROM issues WHERE project_id = ?', (project_id,))
    total = cursor.fetchone()[0]

    cursor.execute('SELECT COUNT(DISTINCT "研发工程师") FROM issues WHERE project_id = ?', (project_id,))
    engineers = cursor.fetchone()[0]

    cursor.execute(
        'SELECT "问题点当前处理进度", COUNT(*) FROM issues WHERE project_id = ? GROUP BY "问题点当前处理进度"',
        (project_id,))
    status_stats = {row[0] or '未填写': row[1] for row in cursor.fetchall()}

    conn.close()
    return jsonify({
        'success': True,
        'total': total,
        'engineers': engineers,
        'status_stats': status_stats
    })


if __name__ == '__main__':
    init_db()
    print('=' * 50)
    print('Issue 跟踪小程序已启动')
    print('访问地址: http://127.0.0.1:5000')
    print(f'管理员密码: {app.config["ADMIN_PASSWORD"]}')
    print('=' * 50)
    app.run(debug=True, host='0.0.0.0', port=5000)
